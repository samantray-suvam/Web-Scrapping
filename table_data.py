from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urljoin
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time

# Launch browser and go to homepage
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()
parent_url = "https://rnbtender.nprocure.com/"
driver.get(parent_url)
print("⏳ Waiting for page to load...")
time.sleep(5)  # Let pop-up auto-close

# Click the anchor with id="tenderInProgress"
try:
    tender_in_progress = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, "tenderInProgress"))
    )
    tender_in_progress.click()
    print("✅ Clicked 'Tender In Progress'")
except Exception as e:
    print(f"❌ Could not click 'Tender In Progress': {e}")
    driver.quit()
    exit()

# Switch to the new tab
WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
driver.switch_to.window(driver.window_handles[-1])
print("🔄 Switched to new tab.")

# Wait for tables to load
WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.TAG_NAME, "table"))
)
child_html = driver.page_source
child_soup = BeautifulSoup(child_html, 'html.parser')
child_tables = child_soup.find_all("table")


#printing the cells

h4_tags = child_soup.find_all("h4")
h4_texts = [h4.get_text(strip=True) for h4 in h4_tags]

all_rows = []
bold_cells = []  # Store (row_idx, col_idx) for <th> cells
h4_rows = []     # Store row indices for <h4> rows

h4_idx = 0  # Index for h4_texts


# adding pdf url
for table_idx, table in enumerate(child_tables):
    if h4_idx < len(h4_texts):
        all_rows.append([h4_texts[h4_idx]])
        h4_rows.append(len(all_rows))  # 1-based index for openpyxl
        h4_idx += 1
    for tr in table.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        row = []
        for col_idx, cell in enumerate(cells):
            # For the last table, check for PDF links in <a> tags
            if table_idx == len(child_tables) - 1 and cell.name == "td":
                a_tag = cell.find("a", href=True)
                if a_tag and a_tag["href"].lower().endswith(".pdf"):
                    link_text = a_tag.get_text(strip=True)
                    pdf_url = urljoin("https://rnbtender.nprocure.com", a_tag["href"])
                    # Insert the PDF URL in the 2nd column
                    if col_idx == 1:
                        row.append(pdf_url)
                    else:
                        row.append(cell.get_text(strip=True))
                else:
                    row.append(cell.get_text(strip=True))
            else:
                row.append(cell.get_text(strip=True))
            if cell.name == "th":
                bold_cells.append((len(all_rows)+1, col_idx+1))  # openpyxl is 1-based
        if row:
            all_rows.append(row)
          

# Save all rows to a single sheet
df = pd.DataFrame(all_rows)
excel_file = "all_tables.xlsx"
df.to_excel(excel_file, index=False, header=False)

# Make <th> cells and <h4> rows bold
wb = load_workbook(excel_file)
ws = wb.active
bold_font = Font(bold=True)
# Bold for <th> cells
for row_idx, col_idx in bold_cells:
    ws.cell(row=row_idx, column=col_idx).font = bold_font
# Bold for <h4> rows (entire row)
for row_idx in h4_rows:
    for cell in ws[row_idx]:
        cell.font = bold_font
    
wb.save(excel_file)

print(f"✅ All tables combined and saved to {excel_file} with bold headers and section titles!")

print(f"🎉 All tables saved to {excel_file}")
driver.quit()